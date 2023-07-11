import argparse

from openpyxl import load_workbook
from typing import Tuple, Optional, Iterable
from pathlib import Path


def get_cells(
    xlsx_filename: Path,
    number_of_columns: int = 10,
    start_row: int = 2
) -> Iterable[
    Tuple[str, str, Optional[str], Optional[str], str, str, str, str, str, str]
]:
    """
    :param: active_worksheet
    :return: генератор возвращает кортеж. Каждый кортеж, соответствует строке таблицы.
    Если у обрабатываемой строки первая ячейка пустая - цикл прерывается. Обработаются только те ячейки,
    которые были прочитаны до прерывания цикла.
    """
    wb = load_workbook(filename=xlsx_filename, read_only=True)
    active_worksheet = wb.active

    for row in active_worksheet.iter_rows(min_row=start_row, max_col=number_of_columns):
        if row[0].value is None:
            break

        current_row_cells: list = []

        for cell in row:
            current_row_cells.append(cell.value)

        yield tuple(current_row_cells)


def get_args():
    parser = argparse.ArgumentParser(
        prog="huntflow_import",
        description="The script interact with huntflow API to upload applicants into it."
    )
    parser.add_argument(
        "--db_path",
        type=Path,
        required=True,
        help="Path to db.",
    )
    argv = parser.parse_args()

    return argv.db_path


def get_xlsx_file_path(directory: Path):
    for file in directory.iterdir():
        if file.suffix == ".xlsx":
            return file
    raise FileNotFoundError
